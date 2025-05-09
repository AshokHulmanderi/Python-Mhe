import openpyxl
import os
from openpyxl import Workbook

def write_to_excel(file_path, message, parsed_message, max_rows=100):
    try:
        # Try to load the existing workbook
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        else:
            # If the file doesn't exist, create a new workbook
            workbook = Workbook()
            sheet = workbook.active
            # Set headers (optional)
            sheet.append(["Message", "Parsed Message"])

        # Check the current number of rows in the sheet
        current_row_count = sheet.max_row

        # If the current row count exceeds the max_rows, overwrite the file
        if current_row_count >= max_rows + 1:  # +1 to account for the header row
            workbook = Workbook()
            sheet = workbook.active
            # Set headers again
            sheet.append(["Message", "Parsed Message"])

        # Write the message and parsed_message to the sheet
        sheet.append([message, str(parsed_message)])  # Convert parsed_message to string if it's a dictionary

        # Save the workbook to the specified file path
        workbook.save(file_path)
        print(f"Data written to {file_path}")

    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")
        

def retrieve_history(row_count: int, loc: str):
    file_path = f"./History_Files/{loc}_Message_History.xlsx"

    # Check if the file exists
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        raise ValueError(f"File '{file_path}' not found.")

    # Read the specified number of rows
    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i == 1:  # Skip the header row
            continue
        if i > row_count + 1:  # Adjust for the skipped header
            break
        rows.append(row)

    # Return the rows as a list of dictionaries
    return {"history": rows}

def clear_history_file(loc: str):
    file_path = f"./History_Files/{loc}_Message_History.xlsx"
   
    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        raise ValueError(f"No history file found to clear for location '{loc}'.")
 
