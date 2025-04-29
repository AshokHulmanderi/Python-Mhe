
import openpyxl
import os

from message_parser import *
from history_tracker import *
from MAWM_Connect import *


# Function to process the appropriate vendor-specific file
def parse_dematic_message(message: str):

    file_path = "./Mapping_Spec/Dematic_MHE.xlsx"

    # Check if the file exists
    if not os.path.exists(file_path):
        return {"error": f"File not found for vendor 'Dematic' - {file_path}"}
    
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)

    # Split the message using the delimiter '^' for Dematic
    parts = message.split('^')
    message_type = parts[6] if len(parts) > 6 else None
    # value_at_position_8 = parts[7] if len(parts) > 7 else None

    # Extract the sheet name from the message using the parser
    sheet_name = parse_sheet_name(message_type)
    
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active  # Default to the first sheet

    # Read the rows from the sheet
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)

    # Extract the delimiter from rows where column1 is 'Field Separator'
    for row in rows:
        if row[0] == 'Field Separator':
            delimiter = row[3]  # Assuming column4 is at index 3
            break
    if not delimiter:
        raise ValueError("Delimiter not found in rows.")

    delimiter2 = None
    # Extract the Custom delimiter from rows where column1 is 'Field Separator'
    for row in rows:
        if row[0] == 'Field Separator' and row[3] != delimiter:
            delimiter2 = row[3]  # Assuming column4 is at index 3
            break
    
    # Validate if delimiter2 exists and is not in the message
    if delimiter2 and delimiter2 not in message: 
        raise ValueError(f"Message does not contain the required custom delimiter '{delimiter2}'.")

    # Extract prefix and suffix from the first and last rows
    prefix = rows[1][3] if len(rows) > 0 and len(rows[1]) > 0 else None
    suffix = rows[-1][3] if len(rows) > 0 and len(rows[-1]) > 0 else None

    # Validate prefix/suffix.
    if not (message.startswith(prefix) and message.endswith(suffix)):
        raise ValueError(f"Message must start with {prefix} and end with {suffix}: {message}")

    original_message = message
    message = message.strip()
    message = message[len(prefix):-len(suffix)].strip()

     # Remove rows where column1 is 'Field Separator'
    rows = [row for row in rows if row[0] != 'Field Separator']

    fields = message.split(delimiter)

    if delimiter2:
        fields = [subfield for field in fields for subfield in field.split(delimiter2)]
        
    # Remove the first two rows and the last row
    rows = rows[2:-1]  # Slicing to exclude the first two rows and the last row

    if len(rows) != len(fields):
       raise ValueError(f"Required Number of values are ({len(rows)}), but received only ({len(fields)}).")

    #Call parse_dematic_message instead of returning the vendor value
    parsed_message = message_Parser(fields, rows)

    # Write the message and parsed_message to an Excel file
    write_to_excel("./History_Files/200_Message_History.xlsx", original_message, parsed_message)

    #Return the parsed message
    return parsed_message


# Function to process the appropriate vendor-specific file
def parse_dematic_divert(inputDivertMessages):

    file_path = "./Mapping_Spec//Dematic_MHE.xlsx"

    # Check if the file exists
    if not os.path.exists(file_path):
        return {"error": f"File not found for vendor 'Dematic'"}
    
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)

    message_type = inputDivertMessages.MessageType if inputDivertMessages.MessageType else "CONTAINERSTATUS"

    # Extract the sheet name from the message using the parser
    sheet_name1 = parse_sheet_name(message_type)
    
    if sheet_name1 and sheet_name1 in workbook.sheetnames:
        sheet1 = workbook[sheet_name1]
    else:
        sheet1 = workbook.active  # Default to the first sheet

    # Extract the sheet name from the message using the parser
    sheet_name2 = parse_sheet_name("CNTR_ERROR")
    
    if sheet_name2 and sheet_name2 in workbook.sheetnames:
        sheet2 = workbook[sheet_name2]
    else:
        sheet2 = workbook[sheet_name2]  # Default to the container divert sheet


    # Read the rows from the sheet
    contdivertrows = []
    for row in sheet1.iter_rows(values_only=True):
        contdivertrows.append(row)

    # Read the rows from the sheet
    contErrorRows = []
    for descrow in sheet2.iter_rows(values_only=True):
        contErrorRows.append(descrow)

    # Remove the first row from contdivertrows
    contdivertrows = contdivertrows[1:]  # Slicing to exclude the first row
    contErrorRows = contErrorRows[1:]  # Slicing to exclude the first row

     # Create a string based on the specified conditions
    result_string = ""
    for row in contdivertrows:
        if row[0] in inputDivertMessages:
            result_string += str(inputDivertMessages[row[0]])  # Use value from inputDivertMessages
        else:
            result_string += str(row[3])  # Default to value from column 1 (index 0)
        result_string += "^"  # Add delimiter after each value

    # Remove the trailing delimiter
    result_string = result_string.rstrip("^")

    error_row_values = [row[0] for row in contErrorRows if len(row) > 0]

    endpointId = "OBPUTAWAY_IB_ENDPOINT_SRC"
    response = invoke_mawm_dei(endpointId, result_string)

    # Write the message and parsed_message to an Excel file
    #write_to_excel("Message_History.xlsx", inputDivertMessages, response)

    return (response)
    #return {"result_string": result_string, "EventResult" : error_row_values }
    #return {"contdivertrows": contdivertrows, "contErrorRows": contErrorRows} 
