#Backup
from typing import Union
from fastapi import FastAPI, UploadFile, File
from pydantic import BaseModel
import openpyxl
import os

class DivertMessage(BaseModel):
    OlpnId: str
    DivertLocation: str

class DivertList(BaseModel):
    Data: List[DivertMessage]

    response = {}
    for divertMessage in divertMessages.Data:
        response[f"{divertMessage.OlpnId}"] = "Diverted to " + divertMessage.DivertLocation
    return Message

# Create a FastAPI instance
app = FastAPI(
    title="Cencora MHE TEST",
    description="This is a sample MHE Dematic Test API",
    version="1.0.0"
)

# Function to read the "MHE Vendor" for a given LOC from Cencora_MHE_LOC.xlsx
def get_mhe_vendor(loc, file_path="Cencora_MHE_LOC.xlsx"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        if row[0] == loc:  # Assuming LOC is in the first column
            return row[2]  # Assuming "MHE Vendor" is in the third column
    return None

# Function to process the appropriate tab in the vendor-specific file
def process_vendor_file(vendor, file_paths, sheet_name=None):
    if vendor not in file_paths:
        return {"error": f"No file found for vendor '{vendor}'"}

    file_path = file_paths[vendor]
    workbook = openpyxl.load_workbook(file_path)

    # If a specific sheet name is provided, use it; otherwise, use the first sheet
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active  # Default to the first sheet

    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)

    return {"vendor": vendor, "sheet_name": sheet.title, "rows": rows}

# Endpoint to validate an XLSX file and process based on LOC
@app.post("/process-loc/")
async def process_loc_endpoint(loc: int, sheet_name: Union[str, None] = None, file: UploadFile = File(...)):
    # Save the uploaded file temporarily
    file_location = f"temp_{file.filename}"
    with open(file_location, "wb") as f:
        f.write(await file.read())

    # Get the MHE Vendor for the given LOC
    mhe_vendor = get_mhe_vendor(loc)
    if not mhe_vendor:
        os.remove(file_location)
        return {"error": f"No MHE Vendor found for LOC {loc}"}

    # Define file paths for vendor-specific files
    vendor_files = {
        "Dematic": "Dematic_MHE.xlsx",
        "Knapp": "Knapp_MHE.xlsx",
        "SSI": "SSI_MHE.xlsx",  # Add other vendor files here
    }

    # Process the appropriate vendor file and sheet
    result = process_vendor_file(mhe_vendor, vendor_files, sheet_name)

    # Delete the temporary file
    os.remove(file_location)

    return result


# Function to validate the XLSX file
def validate_xlsx(file_path):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    errors = []  # List to store validation errors

    # Iterate through rows and columns
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip the header row
        for col_index, value in enumerate(row, start=1):
            if value is not None:
                # Validate the value (e.g., check its size/length)
                if isinstance(value, str) and len(value) > 50:  # Example: Max length = 50
                    errors.append(f"Row {row_index}, Column {col_index}: Value '{value}' exceeds 50 characters.")
                elif isinstance(value, (int, float)) and value > 1000:  # Example: Max numeric value = 1000
                    errors.append(f"Row {row_index}, Column {col_index}: Value '{value}' exceeds the allowed limit.")
            else:
                errors.append(f"Row {row_index}, Column {col_index}: Empty value found.")

    return errors

# Root endpoint
@app.get("/")
def read_root():
    return {"message": "Welcome to the FastAPI application!"}

# Endpoint to validate an XLSX file
@app.post("/validate-xlsx/")
async def validate_xlsx_endpoint(file: UploadFile = File(...)):
    # Save the uploaded file temporarily
    file_location = f"temp_{file.filename}"
    with open(file_location, "wb") as f:
        f.write(await file.read())

    # Call the validate_xlsx function
    errors = validate_xlsx(file_location)

    # Delete the temporary file
    import os
    os.remove(file_location)

    # Return the validation results
    if errors:
        return {"status": "Validation failed", "errors": errors}
    else:
        return {"status": "Validation successful", "message": "No errors found in the file."}
    
    # Endpoint to validate the divertmessages
@app.post("/mhe/Dematic/createDivertMessages")
def dematic_mawm_divert(divertMessages: DivertList):
    response = {}
    for divertMessage in divertMessages.Data:
        response[f"{divertMessage.OlpnId}"] = "Diverted to " + divertMessage.DivertLocation
    return response

